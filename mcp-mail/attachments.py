"""Shared attachment helpers for the mail providers.

Attachments arrive as a list of {"filename", "content_base64"} - the model carries the
bytes from the drive sidecar (via its download_file tool) to here. Small files only: the
bytes pass through the model context, so each is capped. NOTE: sending an attachment to an
EXTERNAL recipient still goes through the per-send consent TOTP (main.py) like any send -
that is the exfiltration guard on "email this document out".
"""

from __future__ import annotations

import base64
import mimetypes
from email.message import EmailMessage

# Mirror the drive sidecar's download cap; per-attachment ceiling.
_ATTACH_MAX_BYTES = 1_000_000


def _decode(att: dict) -> tuple[str, bytes]:
    filename = att.get("filename") or "attachment"
    data = base64.b64decode(att.get("content_base64") or "")
    if len(data) > _ATTACH_MAX_BYTES:
        raise ValueError(f"Attachment '{filename}' too large ({len(data)} bytes > {_ATTACH_MAX_BYTES}).")
    return filename, data


def _guess_type(filename: str) -> tuple[str, str]:
    ctype, _ = mimetypes.guess_type(filename)
    if not ctype:
        return ("application", "octet-stream")
    maintype, _, subtype = ctype.partition("/")
    return (maintype, subtype or "octet-stream")


def add_to_email(mime: EmailMessage, attachments: list[dict] | None) -> None:
    """Attach files to an EmailMessage (Gmail + iCloud, which both build one)."""
    for att in attachments or []:
        filename, data = _decode(att)
        maintype, subtype = _guess_type(filename)
        mime.add_attachment(data, maintype=maintype, subtype=subtype, filename=filename)


def build_xlsx(sheets: list[dict]) -> bytes:
    """Build an .xlsx workbook from sheet DATA (list of {"name","rows"}). Used so the
    model can email a spreadsheet by passing the small data, NOT a giant base64 string
    (LLMs are slow + unreliable at emitting base64 - that broke the model-courier design).
    """
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in sheets or []:
        ws = wb.create_sheet(title=(s.get("name") or "Sheet")[:31])  # Excel caps titles at 31
        for row in s.get("rows", []):
            ws.append(list(row))
    if not wb.worksheets:
        wb.create_sheet(title="Sheet1")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_graph(attachments: list[dict] | None) -> list[dict]:
    """Build Microsoft Graph fileAttachment objects for /me/sendMail (Outlook)."""
    out: list[dict] = []
    for att in attachments or []:
        filename, data = _decode(att)  # validates size; Graph wants the raw base64 back
        out.append({
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": filename,
            "contentBytes": base64.b64encode(data).decode(),
        })
    return out
